import os
from decimal import Decimal
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from SRC.data_handlers import load_users
from SRC.fraud_rules import evaluate_transaction_risk
from SRC.transactions_engine import execute_transaction

# Temporary dictionary acting as our secure memory vault for pending payments
pending_transaction_vault = {}

# Path to our Excel file right inside your project folder hierarchy
EXCEL_LEDGER_PATH = "GuardianShield_Transaction_Ledger.xlsx"

def log_transaction_to_excel(from_user, to_user, amount_str, risk_score):
    """
    Automatically creates or updates an elegant Excel spreadsheet layout
    every single time a transaction successfully clears!
    """
    file_exists = os.path.exists(EXCEL_LEDGER_PATH)
    
    if not file_exists:
        # Create a fresh workbook if it doesn't exist yet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ledger Audit Track"
        ws.views.sheetView[0].showGridLines = True
        
        # Add a nice top banner row
        ws.merge_cells("A1:F1")
        banner = ws["A1"]
        banner.value = "GuardianShield Financial Systems — Live Transaction Audit Ledger"
        banner.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        banner.fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        banner.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add Table Headers
        headers = ["From User ID", "To User ID", "Amount ($)", "Transaction Date", "Execution Time", "Risk Score Metrics"]
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header_title
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[3].height = 22
        start_row = 4
    else:
        wb = openpyxl.load_workbook(EXCEL_LEDGER_PATH)
        ws = wb.active
        start_row = ws.max_row + 1

    # Grab the current real-world timestamp parameters
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    
    # Pack row data matching requirements
    row_values = [from_user, to_user, float(amount_str), date_str, time_str, int(risk_score)]
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Write values into the next available row slot
    for col_idx, val in enumerate(row_values, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = val
        cell.font = Font(name="Segoe UI", size=10, color="2D3748")
        cell.border = thin_border
        
        # Professional adjustments per column type
        if col_idx in [1, 2, 4, 5]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 3:
            cell.number_format = '0.0000000' # Enforces exact banking decimal format display
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx == 6:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            # Apply our warm alert styling block if risk requires attention
            if val >= 31:
                cell.fill = PatternFill(start_color="FEEBC8", end_color="FEEBC8", fill_type="solid")
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="C05621")

    # Give columns some breathing room so names aren't cut off
    for col in ws.columns:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 18

    wb.save(EXCEL_LEDGER_PATH)


def process_payment_request(from_user, to_user, amount_str, hour, device, location):
    """
    Coordinates the validation gate for an incoming payment attempt.
    """
    global pending_transaction_vault
    df_users = load_users()
    
    if from_user not in df_users['user_id'].values:
        return "REJECTED", "Sender account does not exist."
        
    user_profile = df_users[df_users['user_id'] == from_user].iloc[0]
    
    # CRITICAL GATE 1: Safety Check - Instantly reject if account is frozen
    if user_profile['account_status'] == 'LOCKED':
        return "REJECTED", "Transaction denied. This account is currently LOCKED due to security flags."
        
    # CRITICAL GATE 2: Math Check - Instantly reject if funds are insufficient BEFORE calculating risk score
    current_balance = Decimal(str(user_profile['current_balance']))
    requested_amount = Decimal(str(amount_str))
    
    if requested_amount > current_balance:
        missing_amount = requested_amount - current_balance
        return "REJECTED", f"You are out of your money! Try adding ${missing_amount:.7f} more in your account."
        
    # CRITICAL GATE 3: Run quantitative fraud analysis core logic
    risk_score, action, reasons = evaluate_transaction_risk(
        from_user, amount_str, hour, device, location, user_profile
    )
    
    pending_transaction_vault = {
        "from_user_id": from_user,
        "to_user_id": to_user,
        "amount": amount_str,
        "risk_score": risk_score,
        "reasons": reasons
    }
    
    if action == "BLOCK":
        pending_transaction_vault.clear()
        return "REJECTED", f"Blocked by automated fraud rules. Reasons: {', '.join(reasons)}"
        
    elif action == "OTP_CHALLENGE":
        return "OTP_REQUIRED", f"Security Verification Required. Risk Score: {risk_score}"
        
    else:  # APPROVE
        success, message = execute_transaction(from_user, to_user, amount_str)
        if success:
            # AUTO-UPDATE LOGS: Appends row data immediately on instant approvals
            log_transaction_to_excel(from_user, to_user, amount_str, risk_score)
        pending_transaction_vault.clear()
        return "SUCCESS", message


def confirm_otp_and_release_vault(is_otp_correct, screen_amount_str=None):
    """
    Processes the held temporary vault transaction based on OTP validation results.
    """
    global pending_transaction_vault
    
    if not pending_transaction_vault:
        return False, "No active pending transaction found in memory vault."
        
    if is_otp_correct:
        if screen_amount_str and pending_transaction_vault["amount"] != str(screen_amount_str):
            pending_transaction_vault.clear()
            return False, "Security Block: Transaction amount changed mid-session. Vault cleared."

        success, message = execute_transaction(
            pending_transaction_vault["from_user_id"],
            pending_transaction_vault["to_user_id"],
            pending_transaction_vault["amount"]
        )
        
        if success:
            # AUTO-UPDATE LOGS: Appends row data immediately when an OTP challenge is successfully cleared!
            log_transaction_to_excel(
                pending_transaction_vault["from_user_id"],
                pending_transaction_vault["to_user_id"],
                pending_transaction_vault["amount"],
                pending_transaction_vault["risk_score"]
            )
            
        pending_transaction_vault.clear()
        return True, message
    else:
        pending_transaction_vault.clear()
        return False, "Payment failed: Invalid One-Time Password verification code entered."